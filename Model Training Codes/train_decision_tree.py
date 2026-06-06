#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================================
CA-EDR Project — TinyML Decision Tree Classifier
Padma River Water Quality Anomaly Detection
=============================================================================

WHAT THIS SCRIPT DOES:
  Trains and evaluates a TinyML-constrained Decision Tree classifier on the
  Padma River enriched v5 synthetic dataset (Model B — final model).

HOW TO RUN:
  pip install pandas numpy scikit-learn matplotlib seaborn openpyxl
  python train_decision_tree.py

INPUTS:
  padma_synthetic_17280_v5.csv       (enriched dataset)

OUTPUTS:
  results/classification_report_modelB.txt
  results/confusion_matrix_modelB.png
  results/feature_importance_modelB.png
  results/decision_tree_modelB.png
  results/roc_curve_modelB.png
  results/cross_validation_results.csv
  results/full_evaluation_report.txt
  results/training_summary.xlsx

REFERENCES:
  Huang 2025         — stratified sampling, moderate augmentation volume
  Zhao et al. 2023   — stratified sampling ensures subtype coverage
  Wang et al. 2024   — AC-i taxonomy, multi-parameter detection benchmark
  ECR 2023           — anomaly label rule thresholds
=============================================================================
"""

from sklearn.utils import shuffle
from sklearn.metrics import (
    classification_report, confusion_matrix, roc_auc_score,
    roc_curve, accuracy_score, f1_score, precision_score, recall_score
)
from sklearn.model_selection import train_test_split, StratifiedKFold, cross_validate
from sklearn.tree import DecisionTreeClassifier, export_text, plot_tree
import seaborn as sns
import matplotlib.pyplot as plt
import os
import sys
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')


warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

FEATURE_COLS = ["pH", "temperature_C", "TDS_mgl", "DO_mgl"]
LABEL_COL = "label"
SAMPLE_SIZE = 2500
TRAIN_FRAC = 0.70
VAL_FRAC = 0.15
TEST_FRAC = 0.15
RANDOM_SEED = 42

ENRICHED_CSV = "padma_synthetic_17280_v5_with_dates.xlsx"
OUTPUT_DIR = "results"

os.makedirs(OUTPUT_DIR, exist_ok=True)


def load_enriched_dataset(path):
    """Load an enriched dataset from CSV or Excel depending on the file extension."""
    if path.lower().endswith(('.xlsx', '.xls')):
        return pd.read_excel(path, engine='openpyxl')
    if path.lower().endswith('.csv'):
        for encoding in ('utf-8', 'latin1', 'cp1252'):
            try:
                return pd.read_csv(path, encoding=encoding)
            except UnicodeDecodeError:
                continue
        raise UnicodeDecodeError(
            'utf-8', b'', 0, 1,
            f'Unable to decode CSV file using utf-8, latin1 or cp1252: {path}'
        )
    raise ValueError(f"Unsupported dataset format: {path}")

# ─────────────────────────────────────────────────────────────────────────────
# TREE HYPERPARAMETERS
# ─────────────────────────────────────────────────────────────────────────────
# max_depth=4     : ESP32-S3 TinyML constraint — max 31 nodes fits embedded flash
# min_samples_leaf=8  : prevents leaf nodes memorising individual borderline cases
# min_samples_split=15: prevents splitting on CTGAN generation coincidences
# criterion='gini': computationally simpler than entropy — preferred for embedded MCU
# class_weight='balanced': missed anomaly is more dangerous than a false alarm


TREE_PARAMS = dict(
    criterion="gini",
    max_depth=3,
    min_samples_split=15,
    min_samples_leaf=8,
    class_weight="balanced",
    random_state=RANDOM_SEED,
)

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

report_lines = []


def section(title):
    line = "\n" + "=" * 70 + f"\n{title}\n" + "=" * 70
    print(line)
    report_lines.append(line)


def log(text=""):
    print(text)
    report_lines.append(str(text))

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Load the enriched dataset
# ─────────────────────────────────────────────────────────────────────────────


section("STEP 1 — Loading enriched dataset")

try:
    df = load_enriched_dataset(ENRICHED_CSV)
    log(f"  Loaded: {ENRICHED_CSV}  ->  {len(df):,} rows")
except FileNotFoundError:
    sys.exit(
        f"ERROR: '{ENRICHED_CSV}' not found. Run the generation script first.")
except UnicodeDecodeError as e:
    sys.exit(f"ERROR: Could not decode '{ENRICHED_CSV}': {e}")
except ValueError as e:
    sys.exit(f"ERROR: {e}")

n_anom = (df[LABEL_COL] == 1).sum()
n_norm = (df[LABEL_COL] == 0).sum()
log(f"  Anomaly (label=1): {n_anom:,}  ({n_anom/len(df)*100:.2f}%)")
log(f"  Normal  (label=0): {n_norm:,}  ({n_norm/len(df)*100:.2f}%)")

if "anomaly_type" in df.columns:
    log("\n  Anomaly subtype distribution:")
    for atype, cnt in df[df[LABEL_COL] == 1]["anomaly_type"].value_counts().items():
        log(f"    {atype:<30} {cnt:>6}  ({cnt/n_anom*100:.1f}%)")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — Stratified sampling  (Huang 2025, Zhao et al. 2023)
# ─────────────────────────────────────────────────────────────────────────────

section("STEP 2 — Stratified sampling")


def stratified_sample(df, n, label_col=LABEL_COL, seed=RANDOM_SEED):
    n_anom_t = int(n * (df[label_col] == 1).mean())
    n_norm_t = n - n_anom_t
    anom_rows = df[df[label_col] == 1].sample(n=n_anom_t, random_state=seed)
    norm_rows = df[df[label_col] == 0].sample(n=n_norm_t, random_state=seed)
    return pd.concat([anom_rows, norm_rows]).reset_index(drop=True)


log(f"  Sampling {SAMPLE_SIZE} rows (stratified by label)...")
sample = stratified_sample(df, SAMPLE_SIZE)

n1 = (sample[LABEL_COL] == 1).sum()
n0 = (sample[LABEL_COL] == 0).sum()
log(f"  Sample: {len(sample)} rows  —  {n1} anomaly ({n1/len(sample)*100:.1f}%) / {n0} normal ({n0/len(sample)*100:.1f}%)")

if "anomaly_type" in sample.columns:
    log("\n  Subtype coverage after sampling:")
    for atype, cnt in sample[sample[LABEL_COL] == 1]["anomaly_type"].value_counts().items():
        log(f"    {atype:<30} {cnt:>5} rows")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Shuffle and 70/15/15 split
# ─────────────────────────────────────────────────────────────────────────────

section("STEP 3 — Shuffle and 70/15/15 split")

df_s = sample.sample(frac=1, random_state=RANDOM_SEED).reset_index(drop=True)
X = df_s[FEATURE_COLS].values
y = df_s[LABEL_COL].values

X_temp, X_test, y_temp, y_test = train_test_split(
    X, y, test_size=TEST_FRAC, stratify=y, random_state=RANDOM_SEED)

val_rel = VAL_FRAC / (TRAIN_FRAC + VAL_FRAC)
X_train, X_val, y_train, y_val = train_test_split(
    X_temp, y_temp, test_size=val_rel, stratify=y_temp, random_state=RANDOM_SEED)

log(f"  Train : {len(y_train)}  ({sum(y_train == 1)} anomaly / {sum(y_train == 0)} normal)")
log(f"  Val   : {len(y_val)}  ({sum(y_val == 1)} anomaly / {sum(y_val == 0)} normal)")
log(f"  Test  : {len(y_test)}  ({sum(y_test == 1)} anomaly / {sum(y_test == 0)} normal)")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — Train the Decision Tree
# ─────────────────────────────────────────────────────────────────────────────

section("STEP 4 — Training Decision Tree")

log("  Hyperparameters:")
for k, v in TREE_PARAMS.items():
    log(f"    {k:<22} = {v}")

clf = DecisionTreeClassifier(**TREE_PARAMS)
clf.fit(X_train, y_train)

log(f"\n  Trained — depth: {clf.get_depth()}  leaves: {clf.get_n_leaves()}  nodes: {clf.tree_.node_count}")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — Test set evaluation
# ─────────────────────────────────────────────────────────────────────────────

section("STEP 5 — Test set evaluation")

y_pred = clf.predict(X_test)
y_proba = clf.predict_proba(X_test)[:, 1]

acc = accuracy_score(y_test, y_pred)
prec = precision_score(y_test, y_pred, zero_division=0)
rec = recall_score(y_test, y_pred, zero_division=0)
f1 = f1_score(y_test, y_pred, zero_division=0)
auc = roc_auc_score(y_test, y_proba)
cm = confusion_matrix(y_test, y_pred)
cr = classification_report(y_test, y_pred, target_names=["Normal", "Anomaly"])
fi = dict(zip(FEATURE_COLS, clf.feature_importances_))

log(f"  Accuracy           : {acc:.4f}  ({acc*100:.2f}%)")
log(f"  Precision (anomaly): {prec:.4f}")
log(f"  Recall    (anomaly): {rec:.4f}   <- proportion of real anomalies caught")
log(f"  F1        (anomaly): {f1:.4f}")
log(f"  ROC-AUC            : {auc:.4f}")
log(f"\n  Confusion Matrix:")
log(f"    TN={cm[0, 0]}  FP={cm[0, 1]}")
log(f"    FN={cm[1, 0]}  TP={cm[1, 1]}")
log(f"\n  Feature Importance:")
for feat, imp in sorted(fi.items(), key=lambda x: -x[1]):
    bar = "X" * int(imp * 40)
    log(f"    {feat:<16} {imp:.4f}  {bar}")
log(f"\n  Classification Report:\n{cr}")

fname = f"{OUTPUT_DIR}/classification_report_modelB.txt"
with open(fname, "w", encoding="utf-8") as f:
    f.write("Model B — Enriched v5 Dataset\n" + "="*60 + "\n")
    f.write(
        f"Accuracy : {acc:.4f}\nPrecision: {prec:.4f}\nRecall   : {rec:.4f}\n")
    f.write(f"F1       : {f1:.4f}\nROC-AUC  : {auc:.4f}\n\n")
    f.write(
        f"Confusion Matrix:\n  TN={cm[0, 0]}  FP={cm[0, 1]}\n  FN={cm[1, 0]}  TP={cm[1, 1]}\n\n")
    f.write("Feature Importance:\n")
    for feat, imp in sorted(fi.items(), key=lambda x: -x[1]):
        f.write(f"  {feat:<16} {imp:.4f}\n")
    f.write(f"\nClassification Report:\n{cr}\n")
log(f"  Saved: {fname}")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — 5-fold Stratified Cross-Validation
# ─────────────────────────────────────────────────────────────────────────────

section("STEP 6 — 5-fold Stratified Cross-Validation")

log("  Running 5-fold stratified cross-validation...")

skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=RANDOM_SEED)
X_cv, y_cv = shuffle(
    sample[FEATURE_COLS].values, sample[LABEL_COL].values, random_state=RANDOM_SEED)

cv_res = cross_validate(
    DecisionTreeClassifier(**TREE_PARAMS), X_cv, y_cv, cv=skf,
    scoring=["accuracy", "f1", "roc_auc", "precision", "recall"],
    return_train_score=False
)

cv_records = []
for metric in ["accuracy", "f1", "roc_auc", "precision", "recall"]:
    scores = cv_res[f"test_{metric}"]
    log(f"    {metric:<12} mean={scores.mean():.4f}  std={scores.std():.4f}  "
        f"min={scores.min():.4f}  max={scores.max():.4f}")
    cv_records.append({
        "Model": "Model B — Enriched", "Metric": metric,
        "Mean": round(scores.mean(), 4), "Std": round(scores.std(), 4),
        "Min": round(scores.min(), 4), "Max": round(scores.max(), 4),
        "Scores": scores.tolist()
    })

cv_df = pd.DataFrame(cv_records)
cv_df.to_csv(f"{OUTPUT_DIR}/cross_validation_results.csv", index=False)
log(f"\n  Saved: {OUTPUT_DIR}/cross_validation_results.csv")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 7 — Visualisations
# ─────────────────────────────────────────────────────────────────────────────

section("STEP 7 — Generating visualisations")

# 7A: Confusion matrix
log("  7A: Confusion matrix...")
fig, ax = plt.subplots(figsize=(7, 5))
total = cm.sum()
annot = np.array(
    [[f"{cm[i, j]}\n({cm[i, j]/total*100:.1f}%)" for j in range(2)] for i in range(2)])
sns.heatmap(cm, annot=annot, fmt="", ax=ax, cmap="Blues", cbar=False,
            xticklabels=["Predicted Normal", "Predicted Anomaly"],
            yticklabels=["Actual Normal", "Actual Anomaly"],
            linewidths=0.5, linecolor="white")
ax.set_title("Confusion Matrix — Model B (Enriched v5 Dataset)\nPadma River CA-EDR",
             fontsize=12, fontweight="bold")
ax.text(0.5, -0.14,
        f"Accuracy: {acc*100:.2f}%   Recall: {rec*100:.1f}%   F1: {f1:.3f}   AUC: {auc:.3f}",
        transform=ax.transAxes, ha="center", fontsize=9, color="#333333")
plt.tight_layout()
plt.savefig(f"{OUTPUT_DIR}/confusion_matrix_modelB.png",
            dpi=150, bbox_inches="tight")
plt.close()
log(f"  Saved: {OUTPUT_DIR}/confusion_matrix_modelB.png")

# 7B: Feature importance
log("  7B: Feature importance...")
features_display = ["DO (mg/L)", "pH", "TDS (mg/L)", "Temp (C)"]
features_key = ["DO_mgl", "pH", "TDS_mgl", "temperature_C"]
imp_vals = [fi.get(f, 0) for f in features_key]
fig, ax = plt.subplots(figsize=(8, 5))
bars = ax.bar(features_display, imp_vals, color="#2E75B6",
              alpha=0.85, edgecolor="white", linewidth=0.5)
for bar in bars:
    h = bar.get_height()
    if h > 0.01:
        ax.text(bar.get_x() + bar.get_width()/2, h + 0.01,
                f"{h:.3f}", ha="center", va="bottom", fontsize=9, color="#2E75B6")
ax.set_xlabel("Sensor Parameter", fontsize=11)
ax.set_ylabel("Feature Importance (Gini)", fontsize=11)
ax.set_title("Feature Importance — Model B (Enriched v5 Dataset)\nMulti-parameter Anomaly Detection, Padma River CA-EDR",
             fontsize=12, fontweight="bold")
ax.set_ylim(0, 1.05)
ax.yaxis.grid(True, alpha=0.3)
ax.set_axisbelow(True)
plt.tight_layout()
plt.savefig(f"{OUTPUT_DIR}/feature_importance_modelB.png",
            dpi=150, bbox_inches="tight")
plt.close()
log(f"  Saved: {OUTPUT_DIR}/feature_importance_modelB.png")

# 7C: ROC curve
log("  7C: ROC curve...")
fig, ax = plt.subplots(figsize=(8, 6))
fpr, tpr, _ = roc_curve(y_test, y_proba)
ax.plot(fpr, tpr, color="#2E75B6", linewidth=2.2,
        label=f"Model B — Enriched  (AUC = {auc:.4f})")
ax.plot([0, 1], [0, 1], "k--", linewidth=1, alpha=0.4,
        label="Random classifier (AUC = 0.5)")
ax.set_xlabel("False Positive Rate", fontsize=11)
ax.set_ylabel("True Positive Rate (Recall)", fontsize=11)
ax.set_title("ROC Curve — Model B (Enriched v5 Dataset)\nAnomaly Detection, Padma River CA-EDR",
             fontsize=12, fontweight="bold")
ax.legend(fontsize=10, loc="lower right")
ax.set_xlim([0, 1])
ax.set_ylim([0, 1.02])
ax.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig(f"{OUTPUT_DIR}/roc_curve_modelB.png", dpi=150, bbox_inches="tight")
plt.close()
log(f"  Saved: {OUTPUT_DIR}/roc_curve_modelB.png")

# 7D: Decision tree diagram
log("  7D: Decision tree diagram...")
fig, ax = plt.subplots(figsize=(22, 10))
plot_tree(clf, feature_names=["pH", "Temp (C)", "TDS (mg/L)", "DO (mg/L)"],
          class_names=["Normal", "Anomaly"], filled=True, rounded=True,
          impurity=True, proportion=False, ax=ax, fontsize=9)
ax.set_title("Decision Tree — Model B (Enriched v5 Dataset)\nMulti-parameter Anomaly Detection, Padma River CA-EDR",
             fontsize=13, fontweight="bold", pad=15)
plt.tight_layout()
plt.savefig(f"{OUTPUT_DIR}/decision_tree_modelB.png",
            dpi=150, bbox_inches="tight")
plt.close()
log(f"  Saved: {OUTPUT_DIR}/decision_tree_modelB.png")

# 7E: Decision tree text rules
tree_rules = export_text(
    clf, feature_names=["pH", "temperature_C", "TDS_mgl", "DO_mgl"])
rules_path = f"{OUTPUT_DIR}/decision_tree_rules_modelB.txt"
with open(rules_path, "w", encoding="utf-8") as f:
    f.write("Decision Tree Rules — Model B (Enriched v5 Dataset)\n" + "="*60 + "\n")
    f.write(tree_rules)
log(f"  Saved: {rules_path}")
log("\n  Decision Tree Rules:\n" + tree_rules)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 8 — Academic interpretation
# ─────────────────────────────────────────────────────────────────────────────

section("STEP 8 — Academic interpretation")

fi_do = fi.get("DO_mgl", 0)
fi_pH = fi.get("pH", 0)
fi_TDS = fi.get("TDS_mgl", 0)
fn = cm[1, 0]
fp = cm[0, 1]

interp = f"""
  ACCURACY:
    {acc*100:.2f}% on enriched multi-parameter classification.
    Non-trivial accuracy with non-zero FP/FN confirms the model is doing real
    statistical work, not memorising a single-feature rule.

  RECALL (anomaly detection rate): {rec:.4f}
    {'Strong — catches the large majority of real anomalies.' if rec > 0.85 else 'Below 0.85 — review class_weight setting and borderline coverage.'}
    For water quality monitoring, recall is the most critical metric.
    A missed pollution event has direct public health consequences.

  FEATURE IMPORTANCE:
    DO_mgl : {fi_do:.4f}
    pH     : {fi_pH:.4f}
    TDS_mgl: {fi_TDS:.4f}
    {'DO dominance REDUCED — pH and TDS contribute meaningfully. Multi-parameter learning CONFIRMED.' if fi_do < 0.75 else 'DO still dominates — inspect tree diagram for pH/TDS contributions at deeper splits.'}

  CONFUSION MATRIX:
    FP = {fp}  (normal readings flagged as anomaly — false alarms)
    FN = {fn}  (real anomalies missed — the more dangerous error)
    {'Non-zero FP and FN confirmed — model is NOT trivially classifying.' if fp > 0 and fn > 0 else 'Check if FP or FN is zero — may indicate residual trivial classification.'}

  AUC: {auc:.4f}
    {'Excellent discriminability across all classification thresholds.' if auc >= 0.93 else 'Good discriminability.' if auc >= 0.88 else 'AUC below 0.88 — review dataset and hyperparameters.'}
"""
log(interp)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 9 — Save full report and Excel summary
# ─────────────────────────────────────────────────────────────────────────────

section("STEP 9 — Saving full evaluation report")

report_path = f"{OUTPUT_DIR}/full_evaluation_report.txt"
with open(report_path, "w", encoding="utf-8") as f:
    f.write("CA-EDR Decision Tree Evaluation Report — Model B (Enriched v5)\n")
    f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    f.write("=" * 70 + "\n\n")
    f.write("\n".join(report_lines))
    f.write(interp)
log(f"  Saved: {report_path}")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model B Summary"
    alt = PatternFill("solid", start_color="EBF5FB")
    ok = PatternFill("solid", start_color="E2EFDA")
    thin = Side(style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:B1")
    ws["A1"] = "CA-EDR Decision Tree — Model B (Enriched v5 Dataset)"
    ws["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22

    for c, h in enumerate(["Metric", "Model B — Enriched"], 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", start_color="2E75B6")
        cell.alignment = Alignment(horizontal="center")
        cell.border = bdr

    summary_rows = [
        ("Accuracy",            f"{acc*100:.2f}%"),
        ("Precision (anomaly)", f"{prec:.4f}"),
        ("Recall (anomaly)",    f"{rec:.4f}"),
        ("F1 (anomaly)",        f"{f1:.4f}"),
        ("ROC-AUC",             f"{auc:.4f}"),
        ("DO importance",       f"{fi.get('DO_mgl', 0):.4f}"),
        ("pH importance",       f"{fi.get('pH', 0):.4f}"),
        ("TDS importance",      f"{fi.get('TDS_mgl', 0):.4f}"),
        ("Temp importance",     f"{fi.get('temperature_C', 0):.4f}"),
        ("True Negatives",      str(cm[0, 0])),
        ("False Positives",     str(cm[0, 1])),
        ("False Negatives",     str(cm[1, 0])),
        ("True Positives",      str(cm[1, 1])),
        ("Tree depth",          str(clf.get_depth())),
        ("Tree nodes",          str(clf.tree_.node_count)),
    ]
    for r, (metric, value) in enumerate(summary_rows, 3):
        fill = ok if r % 2 == 0 else alt
        for c, val in enumerate([metric, value], 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="center" if c > 1 else "left")
            cell.border = bdr

    ws2 = wb.create_sheet("Cross-Validation")
    for c_name, w in zip(["A", "B", "C", "D", "E", "F"], [25, 16, 12, 12, 12, 12]):
        ws2.column_dimensions[c_name].width = w
    for c, h in enumerate(["Model", "Metric", "Mean", "Std", "Min", "Max"], 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.border = bdr
    for r, row in enumerate(cv_df.itertuples(index=False), 2):
        for c, val in enumerate([row.Model, row.Metric, row.Mean, row.Std, row.Min, row.Max], 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = alt if r % 2 == 0 else ok
            cell.border = bdr

    wb.save(f"{OUTPUT_DIR}/training_summary.xlsx")
    log(f"  Saved: {OUTPUT_DIR}/training_summary.xlsx")
except ImportError:
    log("  openpyxl not installed — skipping Excel output")

# ─────────────────────────────────────────────────────────────────────────────
# FINAL SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

section("TRAINING COMPLETE — QUICK REFERENCE SUMMARY")

log(f"""
  MODEL B (Enriched v5 seed — FINAL MODEL):
    Accuracy  : {acc*100:.2f}%
    Recall    : {rec:.4f}
    F1        : {f1:.4f}
    AUC       : {auc:.4f}
    DO import : {fi.get('DO_mgl', 0):.4f}
    pH import : {fi.get('pH', 0):.4f}
    TDS import: {fi.get('TDS_mgl', 0):.4f}
    FP / FN   : {cm[0, 1]} / {cm[1, 0]}
    Tree depth: {clf.get_depth()}  |  Nodes: {clf.tree_.node_count}

  OUTPUT FILES:
    {OUTPUT_DIR}/confusion_matrix_modelB.png
    {OUTPUT_DIR}/feature_importance_modelB.png
    {OUTPUT_DIR}/roc_curve_modelB.png
    {OUTPUT_DIR}/decision_tree_modelB.png
    {OUTPUT_DIR}/decision_tree_rules_modelB.txt
    {OUTPUT_DIR}/classification_report_modelB.txt
    {OUTPUT_DIR}/cross_validation_results.csv
    {OUTPUT_DIR}/full_evaluation_report.txt
    {OUTPUT_DIR}/training_summary.xlsx
""")
