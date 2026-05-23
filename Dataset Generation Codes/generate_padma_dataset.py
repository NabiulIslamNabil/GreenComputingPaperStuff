#!/usr/bin/env python3
"""
=============================================================================
Padma River CA-EDR Synthetic Dataset Generator
Following Chia et al. (2023) framework — Steps 3 to 5
=============================================================================

Simulation config:
  Nodes    : 8
  Days     : 15
  Interval : 10 minutes → 144 readings/day
  Total    : 8 × 15 × 144 = 17,280 rows

Label rule (ECR 2023, Tofshil-2, pp.53 & 57):
  label = 1  if  DO < 4.0  OR  pH < 6.0
              OR  pH > 8.5 OR  TDS > 1000

Usage:
  pip install sdv pandas numpy scipy openpyxl
  python generate_padma_dataset.py

Input  : seed_dataset_40rows_v2.csv  (must be in the same folder)
Outputs: padma_river_synthetic_17280.csv
         padma_river_synthetic_17280.xlsx
         model_evaluation_metrics.csv
         model_evaluation_metrics.xlsx
=============================================================================
"""

import sys
import warnings
import pandas as pd
import numpy as np
from datetime import date, timedelta
from scipy import stats
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
SEED_CSV      = "seed_dataset_40rows_v2.csv"
NODES         = 8
DAYS          = 15
INTERVAL_MIN  = 10                              # sampling interval in minutes
READINGS_DAY  = 1440 // INTERVAL_MIN           # 144 readings per day
FINAL_ROWS    = NODES * DAYS * READINGS_DAY    # 17,280
EVAL_ROWS     = 500                            # rows for quality evaluation
EPOCHS        = 500                            # for GAN / TVAE models
RANDOM_SEED   = 42
START_DATE    = date(2024, 1, 1)

# Columns the SDV models train on — metadata & label are EXCLUDED
FEATURE_COLS = ["pH", "temperature_C", "TDS_mgl", "DO_mgl"]

# Physical plausibility bounds for clamping (Chia et al. 2023 + cross-paper)
BOUNDS = {
    "pH":            (6.5,  9.0),
    "temperature_C": (18.0, 36.0),
    "TDS_mgl":       (80.0, 300.0),
    "DO_mgl":        (0.5,  10.0),
}

# ECR 2023 label thresholds (Tofshil-2, pp.53 & 57)
def ecr_label(pH, temp, TDS, DO):
    """
    Temperature is retained as a contextual environmental
    monitoring parameter but excluded from direct anomaly labeling.
    """
    if DO < 4.0 or pH < 6.0 or pH > 8.5 or TDS > 1000.0:
        return 1
    return 0


# ─────────────────────────────────────────────────────────────────────────────
# STEP 0 — Load and prepare seed
# ─────────────────────────────────────────────────────────────────────────────
print("=" * 70)
print("STEP 0 — Loading seed dataset")
print("=" * 70)

try:
    seed_full = pd.read_csv(SEED_CSV)
except FileNotFoundError:
    sys.exit(f"ERROR: '{SEED_CSV}' not found. Place it in the same folder.")

print(f"Loaded {len(seed_full)} rows from '{SEED_CSV}'")

# Drop ALL metadata columns — only the 4 physical parameters go to the model.
# Label is intentionally excluded: we re-derive it from ECR 2023 thresholds
# after generation, ensuring labels are always physically consistent.
COLS_TO_DROP = ["seed_row_id", "source", "label", "label_reason"]
seed_train = seed_full.drop(
    columns=[c for c in COLS_TO_DROP if c in seed_full.columns]
).copy()

assert list(seed_train.columns) == FEATURE_COLS, (
    f"Expected columns {FEATURE_COLS}, got {list(seed_train.columns)}"
)

print(f"Training columns: {FEATURE_COLS}")
print(f"\nSeed statistics:")
print(seed_train.describe().round(3).to_string())


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Build SDV metadata
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 70)
print("STEP 1 — Building SDV metadata")
print("=" * 70)

from sdv.metadata import SingleTableMetadata

metadata = SingleTableMetadata()
metadata.detect_from_dataframe(seed_train)
for col in FEATURE_COLS:
    metadata.update_column(col, sdtype="numerical")

print("Metadata configured — all 4 columns marked as numerical.")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — Train all 4 SDV models
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 70)
print("STEP 2 — Training 4 SDV models on 40-row seed")
print("=" * 70)

from sdv.single_table import (
    CTGANSynthesizer,
    TVAESynthesizer,
    CopulaGANSynthesizer,
    GaussianCopulaSynthesizer,
)

model_configs = {
    "CTGAN": CTGANSynthesizer(
        metadata,
        epochs=EPOCHS,
        verbose=False,
    ),
    "TVAE": TVAESynthesizer(
        metadata,
        epochs=EPOCHS,
    ),
    "CopulaGAN": CopulaGANSynthesizer(
        metadata,
        epochs=EPOCHS,
        verbose=False,
    ),
    "GaussianCopula": GaussianCopulaSynthesizer(
        metadata,
    ),
}

trained_models = {}
for name, model in model_configs.items():
    print(f"  Training {name:<20}", end="", flush=True)
    model.fit(seed_train)
    trained_models[name] = model
    print("✓")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Evaluate all models (PCD + KLD + KS) on 500 eval rows
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 70)
print(f"STEP 3 — Evaluating models on {EVAL_ROWS} eval rows (Chia et al. 2023)")
print("=" * 70)


def compute_pcd(real_df, syn_df, cols):
    C_real = real_df[cols].corr().values
    C_syn  = syn_df[cols].corr().values
    return float(np.linalg.norm(C_real - C_syn, "fro"))


def compute_kld_per_col(real_df, syn_df, cols, n_bins=20):
    kld_per_col = {}
    for col in cols:
        r = real_df[col].values
        s = syn_df[col].values
        lo = min(r.min(), s.min())
        hi = max(r.max(), s.max())
        if lo == hi:
            kld_per_col[col] = 0.0
            continue
        bins = np.linspace(lo, hi, n_bins + 1)
        p, _ = np.histogram(r, bins=bins, density=True)
        q, _ = np.histogram(s, bins=bins, density=True)
        p = p + 1e-10;  p /= p.sum()
        q = q + 1e-10;  q /= q.sum()
        kld_per_col[col] = float(np.sum(p * np.log(p / q)))
    return kld_per_col


def compute_ks_per_col(real_df, syn_df, cols):
    ks_per_col = {}
    for col in cols:
        stat, pval = stats.ks_2samp(real_df[col].values, syn_df[col].values)
        ks_per_col[col] = {"D": float(stat), "p": float(pval)}
    return ks_per_col


eval_records = []
eval_generated = {}

for name, model in trained_models.items():
    print(f"\n  [{name}]")
    gen = model.sample(num_rows=EVAL_ROWS)
    eval_generated[name] = gen

    pcd = compute_pcd(seed_train, gen, FEATURE_COLS)

    kld_cols = compute_kld_per_col(seed_train, gen, FEATURE_COLS)
    kld_total = sum(kld_cols.values())
    kld_mean  = kld_total / len(FEATURE_COLS)

    ks_cols   = compute_ks_per_col(seed_train, gen, FEATURE_COLS)
    ks_mean_d = np.mean([v["D"] for v in ks_cols.values()])
    ks_pass   = sum(1 for v in ks_cols.values() if v["p"] > 0.05)

    composite = pcd + kld_mean + ks_mean_d

    print(f"    PCD       = {pcd:.6f}  (lower better)")
    print(f"    KLD_total = {kld_total:.6f}  KLD_mean = {kld_mean:.6f}")
    for col, kv in kld_cols.items():
        print(f"       {col:<16} KLD = {kv:.6f}")
    print(f"    KS_mean_D = {ks_mean_d:.6f}  ({ks_pass}/{len(FEATURE_COLS)} cols pass p>0.05)")
    for col, kv in ks_cols.items():
        flag = "✓" if kv["p"] > 0.05 else "✗"
        print(f"       {col:<16} D={kv['D']:.4f}  p={kv['p']:.4f}  {flag}")
    print(f"    COMPOSITE = {composite:.6f}")

    eval_records.append({
        "Model":        name,
        "PCD":          round(pcd, 6),
        "KLD_pH":       round(kld_cols["pH"], 6),
        "KLD_temp":     round(kld_cols["temperature_C"], 6),
        "KLD_TDS":      round(kld_cols["TDS_mgl"], 6),
        "KLD_DO":       round(kld_cols["DO_mgl"], 6),
        "KLD_total":    round(kld_total, 6),
        "KLD_mean":     round(kld_mean, 6),
        "KS_pH_D":      round(ks_cols["pH"]["D"], 6),
        "KS_temp_D":    round(ks_cols["temperature_C"]["D"], 6),
        "KS_TDS_D":     round(ks_cols["TDS_mgl"]["D"], 6),
        "KS_DO_D":      round(ks_cols["DO_mgl"]["D"], 6),
        "KS_mean_D":    round(ks_mean_d, 6),
        "KS_pass_cols": ks_pass,
        "Composite":    round(composite, 6),
    })

metrics_df = pd.DataFrame(eval_records).set_index("Model").sort_values("Composite")

print("\n" + "─" * 70)
print("MODEL RANKING (lowest Composite wins):")
print(metrics_df[["PCD", "KLD_mean", "KS_mean_D", "Composite"]].to_string())
print("─" * 70)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — Select winner
# ─────────────────────────────────────────────────────────────────────────────
winner_name  = metrics_df.index[0]
winner_model = trained_models[winner_name]
winner_row   = metrics_df.loc[winner_name]

print(f"\nWINNER: {winner_name}")
print(f"  PCD={winner_row['PCD']:.6f}, "
      f"KLD_mean={winner_row['KLD_mean']:.6f}, "
      f"KS_mean_D={winner_row['KS_mean_D']:.6f}, "
      f"Composite={winner_row['Composite']:.6f}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — Generate final 17,280 rows
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 70)
print(f"STEP 5 — Generating final {FINAL_ROWS} rows with {winner_name}")
print(f"         8 nodes × 15 days × {READINGS_DAY} intervals (every {INTERVAL_MIN} min)")
print("=" * 70)

synthetic_raw = winner_model.sample(num_rows=FINAL_ROWS)

# Clamp to physical plausibility bounds
print("\nPhysical bounds clamping:")
clamped_rows = set()
for col, (lo, hi) in BOUNDS.items():
    before   = synthetic_raw[col].copy()
    synthetic_raw[col] = synthetic_raw[col].clip(lower=lo, upper=hi)
    changed  = (synthetic_raw[col] != before)
    n_changed = changed.sum()
    clamped_rows.update(changed[changed].index.tolist())
    if n_changed > 0:
        print(f"  {col:<16} {n_changed:>5} values clamped  "
              f"(range was [{before.min():.3f}, {before.max():.3f}], "
              f"bound=[{lo},{hi}])")
    else:
        print(f"  {col:<16} no clamping needed")

n_clamped      = len(clamped_rows)
pct_no_clamp   = (1 - n_clamped / FINAL_ROWS) * 100
print(f"\n  Rows needing ANY clamping : {n_clamped} / {FINAL_ROWS}")
print(f"  Rows with NO clamping     : {FINAL_ROWS - n_clamped} ({pct_no_clamp:.2f}%)")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — Re-apply ECR 2023 label rule
# ─────────────────────────────────────────────────────────────────────────────
print("\nApplying ECR 2023 label rule ...")
labels = [
    ecr_label(
        synthetic_raw.loc[i, "pH"],
        synthetic_raw.loc[i, "temperature_C"],
        synthetic_raw.loc[i, "TDS_mgl"],
        synthetic_raw.loc[i, "DO_mgl"],
    )
    for i in synthetic_raw.index
]
synthetic_raw["label"] = labels
anomaly_count  = sum(labels)
anomaly_pct    = anomaly_count / FINAL_ROWS * 100
print(f"  Anomaly rows (label=1): {anomaly_count} ({anomaly_pct:.2f}%)")
print(f"  Normal  rows (label=0): {FINAL_ROWS - anomaly_count} ({100-anomaly_pct:.2f}%)")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 7 — Assign temporal and spatial structure
# ─────────────────────────────────────────────────────────────────────────────
print("\nAssigning node_id, date, day_of_simulation, time_of_day_min ...")

node_ids      = []
dates         = []
days_of_sim   = []
times_of_day  = []

for node in range(1, NODES + 1):
    for day in range(1, DAYS + 1):
        for interval in range(READINGS_DAY):
            node_ids.append(f"NODE_{node:02d}")
            dates.append(str(START_DATE + timedelta(days=day - 1)))
            days_of_sim.append(day)
            times_of_day.append(interval * INTERVAL_MIN)

final_df = pd.DataFrame({
    "date":             dates,
    "node_id":          node_ids,
    "day_of_simulation": days_of_sim,
    "time_of_day_min":  times_of_day,
    "pH":               synthetic_raw["pH"].values.round(2),
    "temperature_C":    synthetic_raw["temperature_C"].values.round(2),
    "TDS_mgl":          synthetic_raw["TDS_mgl"].values.round(2),
    "DO_mgl":           synthetic_raw["DO_mgl"].values.round(2),
    "label":            synthetic_raw["label"].values,
})

print(f"\nFinal dataset shape : {final_df.shape}")
print(f"\nParameter ranges after generation + clamping:")
print(final_df[FEATURE_COLS].describe().round(3).to_string())


# ─────────────────────────────────────────────────────────────────────────────
# STEP 8 — Save outputs
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 70)
print("STEP 8 — Saving outputs")
print("=" * 70)

# ── CSV outputs ──────────────────────────────────────────────────────────────
final_csv   = "padma_river_synthetic_17280.csv"
metrics_csv = "model_evaluation_metrics.csv"

final_df.to_csv(final_csv, index=False)
metrics_df.reset_index().to_csv(metrics_csv, index=False)
print(f"  Saved: {final_csv}")
print(f"  Saved: {metrics_csv}")

# ── Excel outputs ─────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _header_style(cell, bg="2C3E50"):
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="AAAAAA")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _data_style(cell, fill_hex=None):
        thin = Side(style="thin", color="CCCCCC")
        cell.font      = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        if fill_hex:
            cell.fill = PatternFill("solid", start_color=fill_hex)

    # ── Metrics workbook ─────────────────────────────────────────────────────
    wb_m = openpyxl.Workbook()
    ws_m = wb_m.active
    ws_m.title = "Model Evaluation"

    # Title
    ws_m.merge_cells("A1:P1")
    ws_m["A1"] = (f"SDV Model Evaluation — PCD + KLD + KS  "
                  f"(Chia et al. 2023)  |  Seed: 40 rows  |  Eval: {EVAL_ROWS} rows")
    ws_m["A1"].font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws_m["A1"].fill      = PatternFill("solid", start_color="1A5276")
    ws_m["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_m.row_dimensions[1].height = 20

    m_headers = [
        "Model", "PCD", "KLD_pH", "KLD_Temp", "KLD_TDS", "KLD_DO",
        "KLD_total", "KLD_mean", "KS_pH_D", "KS_Temp_D", "KS_TDS_D",
        "KS_DO_D", "KS_mean_D", "KS_pass_cols", "Composite", "Rank"
    ]
    m_widths = [16,10,10,10,10,10,10,10,10,10,10,10,10,12,12,8]
    for c, (h, w) in enumerate(zip(m_headers, m_widths), 1):
        cell = ws_m.cell(row=2, column=c, value=h)
        _header_style(cell)
        ws_m.column_dimensions[get_column_letter(c)].width = w
    ws_m.row_dimensions[2].height = 18

    metrics_sorted = metrics_df.reset_index()
    winner_fill = "D5F5E3"
    other_fill  = "EBF5FB"
    for r_idx, row in metrics_sorted.iterrows():
        excel_row = r_idx + 3
        fill = winner_fill if r_idx == 0 else other_fill
        values = [
            row["Model"], row["PCD"], row["KLD_pH"], row["KLD_temp"],
            row["KLD_TDS"], row["KLD_DO"], row["KLD_total"], row["KLD_mean"],
            row["KS_pH_D"], row["KS_temp_D"], row["KS_TDS_D"], row["KS_DO_D"],
            row["KS_mean_D"], row["KS_pass_cols"], row["Composite"], r_idx + 1
        ]
        for c_idx, val in enumerate(values, 1):
            cell = ws_m.cell(row=excel_row, column=c_idx, value=val)
            _data_style(cell, fill)
            if isinstance(val, float):
                cell.number_format = "0.000000"
        if r_idx == 0:
            ws_m.cell(row=excel_row, column=1).font = Font(name="Arial", bold=True, size=9)

    # Winner annotation
    ws_m.cell(row=len(metrics_sorted)+4, column=1,
               value=f"Winner: {winner_name}").font = Font(bold=True, name="Arial")
    
    # === FIXED LABEL RULE ===
    ws_m.cell(row=len(metrics_sorted)+5, column=1,
               value="Label rule (ECR 2023, Tofshil-2): "
                     "label=1 if DO<4.0 OR pH<6.0 OR pH>8.5 OR TDS>1000"
               ).font = Font(name="Arial", size=9, italic=True)

    wb_m.save("model_evaluation_metrics.xlsx")
    print("  Saved: model_evaluation_metrics.xlsx")

    # ── Final dataset workbook ───────────────────────────────────────────────
    wb_f = openpyxl.Workbook()

    # Summary sheet
    ws_sum = wb_f.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 45

    ws_sum.merge_cells("A1:B1")
    ws_sum["A1"] = "Padma River Synthetic Dataset — Generation Summary"
    ws_sum["A1"].font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws_sum["A1"].fill      = PatternFill("solid", start_color="1A5276")
    ws_sum["A1"].alignment = Alignment(horizontal="center")

    summary_rows = [
        ("Winning model",        winner_name),
        ("PCD",                  f"{winner_row['PCD']:.6f}"),
        ("KLD (mean)",           f"{winner_row['KLD_mean']:.6f}"),
        ("KS mean D-statistic",  f"{winner_row['KS_mean_D']:.6f}"),
        ("Composite score",      f"{winner_row['Composite']:.6f}"),
        ("", ""),
        ("Nodes",                str(NODES)),
        ("Days",                 str(DAYS)),
        ("Interval (min)",       str(INTERVAL_MIN)),
        ("Readings/day",         str(READINGS_DAY)),
        ("Total rows",           f"{FINAL_ROWS:,}"),
        ("", ""),
        ("Anomaly rows (1)",     f"{anomaly_count} ({anomaly_pct:.2f}%)"),
        ("Normal rows (0)",      f"{FINAL_ROWS-anomaly_count} ({100-anomaly_pct:.2f}%)"),
        ("", ""),
        ("Rows clamped",         f"{n_clamped} ({100-pct_no_clamp:.2f}%)"),
        ("Rows NOT clamped",     f"{FINAL_ROWS-n_clamped} ({pct_no_clamp:.2f}%)"),
        ("", ""),
        ("pH range",             f"{final_df['pH'].min():.2f} – {final_df['pH'].max():.2f}"),
        ("Temp range (°C)",      f"{final_df['temperature_C'].min():.2f} – {final_df['temperature_C'].max():.2f}"),
        ("TDS range (mg/L)",     f"{final_df['TDS_mgl'].min():.2f} – {final_df['TDS_mgl'].max():.2f}"),
        ("DO range (mg/L)",      f"{final_df['DO_mgl'].min():.2f} – {final_df['DO_mgl'].max():.2f}"),
        ("", ""),
        ("Start date",           str(START_DATE)),
        ("Label rule source",    "ECR 2023, Tofshil-2, pp.53 & 57"),
    ]
    for r_i, (k, v) in enumerate(summary_rows, 2):
        ws_sum.cell(row=r_i, column=1, value=k).font = Font(name="Arial", bold=bool(k), size=10)
        ws_sum.cell(row=r_i, column=2, value=v).font = Font(name="Arial", size=10)

    # Data sheet (first 5000 rows)
    ws_data = wb_f.create_sheet("Data (first 5000 rows)")
    data_headers = list(final_df.columns)
    data_widths   = [12, 10, 18, 16, 8, 14, 11, 10, 8]
    for c_i, (h, w) in enumerate(zip(data_headers, data_widths), 1):
        cell = ws_data.cell(row=1, column=c_i, value=h)
        _header_style(cell)
        ws_data.column_dimensions[get_column_letter(c_i)].width = w

    anomaly_fill_d = PatternFill("solid", start_color="FADBD8")
    normal_fill_d  = PatternFill("solid", start_color="FFFFFF")
    for r_i, (_, row_data) in enumerate(final_df.head(5000).iterrows(), 2):
        fill = anomaly_fill_d if row_data["label"] == 1 else normal_fill_d
        for c_i, val in enumerate(row_data.values, 1):
            cell = ws_data.cell(row=r_i, column=c_i, value=val)
            _data_style(cell, None)
            if row_data["label"] == 1:
                cell.fill = anomaly_fill_d

    ws_data.freeze_panes = "A2"

    wb_f.save("padma_river_synthetic_17280.xlsx")
    print("  Saved: padma_river_synthetic_17280.xlsx")

except ImportError:
    print("  openpyxl not installed — skipping Excel output (CSV is still saved)")

# ─────────────────────────────────────────────────────────────────────────────
# FINAL SUMMARY
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 70)
print("GENERATION COMPLETE")
print("=" * 70)
print(f"  Winning model   : {winner_name}")
print(f"  PCD             : {winner_row['PCD']:.6f}")
print(f"  KLD (mean)      : {winner_row['KLD_mean']:.6f}")
print(f"  KS mean D       : {winner_row['KS_mean_D']:.6f}")
print(f"  Composite score : {winner_row['Composite']:.6f}")
print(f"")
print(f"  Final rows      : {FINAL_ROWS:,}")
print(f"  Anomaly (1)     : {anomaly_count} ({anomaly_pct:.2f}%)")
print(f"  Normal  (0)     : {FINAL_ROWS - anomaly_count} ({100 - anomaly_pct:.2f}%)")
print(f"  Rows no-clamp   : {FINAL_ROWS - n_clamped} ({pct_no_clamp:.2f}%)")
print(f"")
print(f"  Output files:")
print(f"    padma_river_synthetic_17280.csv")
print(f"    padma_river_synthetic_17280.xlsx")
print(f"    model_evaluation_metrics.csv")
print(f"    model_evaluation_metrics.xlsx")
print("=" * 70)